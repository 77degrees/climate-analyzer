"""
Import historical climate data from multiple sources into Climate Analyzer DB.

Sources:
  1. SwitchBot CSV exports (iCloud) — Jan 1-13, 2026, per-minute
  2. Home Assistant history CSVs (Desktop-Backup) — Sep 2023 - Jul 2025
  3. Excel exports (Desktop-Backup) — Sep 2023 - Oct 2025

Usage:
  python import_historical.py [--db PATH] [--dry-run]
"""

import argparse
import csv
import os
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

DEFAULT_DB = str(Path(__file__).resolve().parent.parent / "data" / "climate.db")

ICLOUD_CSV_DIR = r"D:\Users\kchri\iCloudDrive\Data-CSV"
BACKUP_DIR = r"C:\Users\kchri\OneDrive\Desktop\Desktop-Backup"

# SwitchBot CSV filename → (temp_sensor_id, humidity_sensor_id)
SWITCHBOT_MAP = {
    "Indoor Meter (Bedroom)_data.csv": (1, 2),
    "Indoor Meter (Bedroom)_data 2.csv": (1, 2),
    "Indoor Meter (Kyle\u2019s Office)_data.csv": (5, 6),
    "Outdoor Meter (Patio)_data.csv": (3, 4),
    "Outdoor Meter (Patio)_data 2.csv": (3, 4),
}

# HA entity aliases — renamed entities that map to the same sensor
ENTITY_ALIASES = {
    "sensor.co2_sensor_humidity": "sensor.kitchen_co2_sensor_humidity",
    "sensor.co2_sensor_temperature": "sensor.kitchen_co2_sensor_temperature",
    "climate.ena_s_eight_sleep_side": "climate.ena_s_eight_sleep_side_climate",
}

# HA history CSV files
HA_CSV_FILES = [
    os.path.join(BACKUP_DIR, "history (4).csv"),
    os.path.join(BACKUP_DIR, "history (5).csv"),
]

# Excel files (Indoor/Outdoor temps + humidity + climate data)
EXCEL_FILES = [
    os.path.join(BACKUP_DIR, "Indoor Temps 1.xlsx"),
    os.path.join(BACKUP_DIR, "Indoor Temps 2.xlsx"),
    os.path.join(BACKUP_DIR, "Indoor Humidity 1.xlsx"),
    os.path.join(BACKUP_DIR, "Indoor Humidity 2.xlsx"),
    os.path.join(BACKUP_DIR, "Outside Temps.xlsx"),
    os.path.join(BACKUP_DIR, "Outside Humidity.xlsx"),
    os.path.join(BACKUP_DIR, "Climate Data.xlsx"),
]

# Entities to skip (not useful for climate analysis)
SKIP_ENTITIES = {
    "automation.update_nest_temperatures",
    "switch.bad_nest_pre_release",
    "update.bad_nest_update",
    "binary_sensor.both_onlinestatus_b0e9fe001a2c",
    "select.upstairs_temperature_display_units",
    "select.downstairs_temperature_display_units",
    "number.bedroom_humidifier_mist_level",
    "binary_sensor.bedroom_humidifier_low_water",
    "binary_sensor.bedroom_humidifier_water_tank_lifted",
    "select.bedroom_humidifier_night_light_level",
    "switch.bedroom_humidifier_display",
    "humidifier.bedroom_humidifier",
}

BATCH_SIZE = 5000

# Climate entity domains that have HVAC fields
CLIMATE_DOMAINS = {"climate"}


def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

def get_entity_map(conn):
    """Build entity_id → sensor_id mapping from DB."""
    cur = conn.execute("SELECT id, entity_id FROM sensors")
    mapping = {}
    for row in cur.fetchall():
        mapping[row[1]] = row[0]
    # Add aliases
    for alias, real in ENTITY_ALIASES.items():
        if real in mapping:
            mapping[alias] = mapping[real]
    return mapping


def ensure_unique_index(conn):
    """Add unique constraint for deduplication."""
    conn.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_readings_sensor_ts_unique "
        "ON readings(sensor_id, timestamp)"
    )
    conn.commit()


def insert_readings(conn, rows, source_name):
    """Batch insert readings with INSERT OR IGNORE for deduplication."""
    inserted = 0
    skipped = 0
    for i in range(0, len(rows), BATCH_SIZE):
        batch = rows[i : i + BATCH_SIZE]
        cur = conn.executemany(
            "INSERT OR IGNORE INTO readings "
            "(sensor_id, timestamp, value, hvac_action, hvac_mode, setpoint_heat, setpoint_cool, fan_mode) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            batch,
        )
        inserted += cur.rowcount
        skipped += len(batch) - cur.rowcount
    conn.commit()
    log(f"  {source_name}: {inserted:,} inserted, {skipped:,} duplicates skipped")
    return inserted


# ---------------------------------------------------------------------------
# Source 1: SwitchBot CSVs
# ---------------------------------------------------------------------------

def parse_switchbot_ts(ts_str):
    """Parse 'Jan 01, 2026 12:00:32 AM' → datetime UTC."""
    return datetime.strptime(ts_str, "%b %d, %Y %I:%M:%S %p").replace(
        tzinfo=timezone.utc
    )


def import_switchbot_csvs(conn, entity_map, dry_run=False):
    """Import SwitchBot CSV exports from iCloud."""
    log("=== Importing SwitchBot CSVs ===")

    if not os.path.isdir(ICLOUD_CSV_DIR):
        log(f"  iCloud directory not found: {ICLOUD_CSV_DIR}")
        return

    total = 0
    for filename, (temp_sid, hum_sid) in SWITCHBOT_MAP.items():
        filepath = os.path.join(ICLOUD_CSV_DIR, filename)
        if not os.path.isfile(filepath):
            log(f"  Skipping (not found): {filename}")
            continue

        rows = []
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            next(reader)  # skip header
            for row in reader:
                try:
                    ts = parse_switchbot_ts(row[0])
                    ts_str = ts.strftime("%Y-%m-%d %H:%M:%S")
                    temp = float(row[1])
                    humidity = float(row[2])

                    # Temperature reading
                    rows.append((temp_sid, ts_str, temp, None, None, None, None, None))
                    # Humidity reading
                    rows.append((hum_sid, ts_str, humidity, None, None, None, None, None))
                except (ValueError, IndexError):
                    continue

        if not dry_run:
            total += insert_readings(conn, rows, filename)
        else:
            log(f"  {filename}: {len(rows):,} rows (dry run)")

    log(f"  SwitchBot total: {total:,} readings inserted")


# ---------------------------------------------------------------------------
# Source 2: HA History CSVs
# ---------------------------------------------------------------------------

def parse_iso_ts(ts_str):
    """Parse ISO 8601 timestamp → datetime string."""
    if not ts_str:
        return None
    # Handle 2023-09-29T23:00:00.000Z
    ts_str = ts_str.replace("Z", "+00:00")
    try:
        dt = datetime.fromisoformat(ts_str)
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None


def safe_float(val):
    """Parse float, return None if invalid."""
    if not val or val in ("unavailable", "unknown", ""):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def import_ha_csvs(conn, entity_map, dry_run=False):
    """Import HA history CSV exports."""
    log("=== Importing HA History CSVs ===")

    total = 0
    skipped_entities = set()

    for filepath in HA_CSV_FILES:
        if not os.path.isfile(filepath):
            log(f"  Skipping (not found): {filepath}")
            continue

        fname = os.path.basename(filepath)
        log(f"  Processing {fname}...")

        rows = []
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader)

            # Build column index
            cols = {h: i for i, h in enumerate(header)}

            for row in reader:
                entity_id = row[cols.get("entity_id", 0)] if len(row) > 0 else ""

                # Skip non-climate entities
                if entity_id in SKIP_ENTITIES:
                    continue

                # Resolve alias
                entity_id = ENTITY_ALIASES.get(entity_id, entity_id)

                sensor_id = entity_map.get(entity_id)
                if sensor_id is None:
                    skipped_entities.add(entity_id)
                    continue

                # Parse timestamp
                ts_raw = row[cols.get("last_changed", 2)] if len(row) > 2 else ""
                ts = parse_iso_ts(ts_raw)
                if not ts:
                    continue

                domain = entity_id.split(".")[0]

                if domain in CLIMATE_DOMAINS:
                    # Climate entity — extract current_temperature, hvac_action, setpoints
                    current_temp = safe_float(
                        row[cols["current_temperature"]] if "current_temperature" in cols and len(row) > cols["current_temperature"] else None
                    )
                    hvac_action = (
                        row[cols["hvac_action"]] if "hvac_action" in cols and len(row) > cols["hvac_action"] else None
                    )
                    if hvac_action in ("", "unavailable", "unknown"):
                        hvac_action = None
                    # Also check "action" column (history (5) has both)
                    if not hvac_action and "action" in cols and len(row) > cols["action"]:
                        act = row[cols["action"]]
                        if act and act not in ("", "unavailable", "unknown"):
                            hvac_action = act

                    hvac_mode = row[cols.get("state", 1)] if len(row) > 1 else None
                    if hvac_mode in ("unavailable", "unknown", ""):
                        hvac_mode = None

                    setpoint_heat = safe_float(
                        row[cols["temperature"]] if "temperature" in cols and len(row) > cols["temperature"] else None
                    )
                    target_high = safe_float(
                        row[cols["target_temp_high"]] if "target_temp_high" in cols and len(row) > cols["target_temp_high"] else None
                    )
                    target_low = safe_float(
                        row[cols["target_temp_low"]] if "target_temp_low" in cols and len(row) > cols["target_temp_low"] else None
                    )
                    setpoint_cool = target_high if target_high else None
                    if target_low and not setpoint_heat:
                        setpoint_heat = target_low

                    rows.append((
                        sensor_id, ts, current_temp,
                        hvac_action, hvac_mode, setpoint_heat, setpoint_cool, None,
                    ))
                else:
                    # Regular sensor — state is the value
                    value = safe_float(row[cols.get("state", 1)] if len(row) > 1 else None)
                    if value is None:
                        continue
                    rows.append((sensor_id, ts, value, None, None, None, None, None))

        if not dry_run:
            total += insert_readings(conn, rows, fname)
        else:
            log(f"  {fname}: {len(rows):,} rows (dry run)")

    if skipped_entities:
        log(f"  Skipped {len(skipped_entities)} unknown entities: {sorted(skipped_entities)[:10]}...")

    log(f"  HA CSV total: {total:,} readings inserted")


# ---------------------------------------------------------------------------
# Source 3: Excel files
# ---------------------------------------------------------------------------

def parse_excel_ts(ts_val):
    """Parse Excel timestamp (could be datetime or string) → datetime string."""
    if ts_val is None:
        return None
    if isinstance(ts_val, datetime):
        return ts_val.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(ts_val, str):
        # Try common formats: "9/29/2025 4:12 AM", "10/13/2025 4:37 PM"
        for fmt in ["%m/%d/%Y %I:%M %p", "%m/%d/%Y %I:%M:%S %p", "%Y-%m-%d %H:%M:%S"]:
            try:
                return datetime.strptime(ts_val, fmt).strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
    return None


def import_excel_files(conn, entity_map, dry_run=False):
    """Import Excel exports from Desktop-Backup."""
    log("=== Importing Excel Files ===")

    try:
        import openpyxl
    except ImportError:
        log("  openpyxl not installed. Run: pip install openpyxl")
        return

    total = 0
    skipped_entities = set()

    for filepath in EXCEL_FILES:
        if not os.path.isfile(filepath):
            log(f"  Skipping (not found): {filepath}")
            continue

        fname = os.path.basename(filepath)
        log(f"  Processing {fname}...")

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        rows = []
        header = None
        row_count = 0

        for excel_row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(h).lower() if h else "" for h in excel_row]
                cols = {h: i for i, h in enumerate(header)}
                continue

            row_count += 1
            raw = list(excel_row)

            entity_id = str(raw[cols.get("entity_id", 1)]) if len(raw) > 1 else ""
            if entity_id in SKIP_ENTITIES:
                continue
            entity_id = ENTITY_ALIASES.get(entity_id, entity_id)

            sensor_id = entity_map.get(entity_id)
            if sensor_id is None:
                skipped_entities.add(entity_id)
                continue

            ts = parse_excel_ts(raw[cols.get("last_changed", 0)])
            if not ts:
                continue

            domain = entity_id.split(".")[0]

            if domain in CLIMATE_DOMAINS and "current_temperature" in cols:
                current_temp = safe_float(raw[cols["current_temperature"]] if len(raw) > cols["current_temperature"] else None)
                hvac_action = str(raw[cols["hvac_action"]]) if "hvac_action" in cols and len(raw) > cols["hvac_action"] and raw[cols["hvac_action"]] else None
                if hvac_action in ("", "None", "unavailable", "unknown"):
                    hvac_action = None
                hvac_mode = str(raw[cols.get("state", 2)]) if len(raw) > 2 and raw[cols.get("state", 2)] else None
                if hvac_mode in ("None", "unavailable", "unknown", ""):
                    hvac_mode = None
                setpoint = safe_float(raw[cols["temperature"]] if "temperature" in cols and len(raw) > cols["temperature"] else None)

                rows.append((sensor_id, ts, current_temp, hvac_action, hvac_mode, setpoint, None, None))
            else:
                value = safe_float(raw[cols.get("state", 2)] if len(raw) > 2 else None)
                if value is None:
                    continue
                rows.append((sensor_id, ts, value, None, None, None, None, None))

            # Periodic batch insert to manage memory
            if len(rows) >= 50000:
                if not dry_run:
                    insert_readings(conn, rows, f"{fname} (batch)")
                rows = []

        wb.close()

        if rows and not dry_run:
            total += insert_readings(conn, rows, fname)
        elif dry_run:
            log(f"  {fname}: {row_count:,} data rows (dry run)")

    if skipped_entities:
        log(f"  Skipped {len(skipped_entities)} unknown entities: {sorted(skipped_entities)[:10]}...")

    log(f"  Excel total: {total:,} readings inserted")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Import historical climate data")
    parser.add_argument("--db", default=DEFAULT_DB, help="Path to climate.db")
    parser.add_argument("--dry-run", action="store_true", help="Don't actually insert")
    parser.add_argument("--skip-excel", action="store_true", help="Skip Excel imports (slow)")
    parser.add_argument("--only", choices=["switchbot", "ha-csv", "excel"], help="Import only one source")
    args = parser.parse_args()

    if not os.path.isfile(args.db):
        log(f"Database not found: {args.db}")
        sys.exit(1)

    log(f"Database: {args.db}")
    conn = sqlite3.connect(args.db)

    # Get initial count
    before = conn.execute("SELECT COUNT(*) FROM readings").fetchone()[0]
    log(f"Readings before import: {before:,}")

    # Build entity mapping
    entity_map = get_entity_map(conn)
    log(f"Loaded {len(entity_map)} entity -> sensor_id mappings")

    # Add unique constraint for deduplication
    log("Ensuring unique index on (sensor_id, timestamp)...")
    ensure_unique_index(conn)

    # Import sources
    if not args.only or args.only == "switchbot":
        import_switchbot_csvs(conn, entity_map, dry_run=args.dry_run)

    if not args.only or args.only == "ha-csv":
        import_ha_csvs(conn, entity_map, dry_run=args.dry_run)

    if (not args.only or args.only == "excel") and not args.skip_excel:
        import_excel_files(conn, entity_map, dry_run=args.dry_run)

    # Final count
    after = conn.execute("SELECT COUNT(*) FROM readings").fetchone()[0]
    log(f"\nReadings after import: {after:,} (+{after - before:,} new)")

    # Date range
    row = conn.execute("SELECT MIN(timestamp), MAX(timestamp) FROM readings").fetchone()
    log(f"Date range: {row[0]} to {row[1]}")

    conn.close()
    log("Done!")


if __name__ == "__main__":
    main()
